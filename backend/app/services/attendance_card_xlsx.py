"""
پارس فایل XLSX «فیش کارکرد پرسنل» — همان ساختار ستونی که ابزار HTML مرجع
(تولیدکننده کارت‌های چاپی «فیش کارکرد») استفاده می‌کند:

ستون‌ها (۰-پایه، یعنی A=0):
    B(1)=کد پرسنلی، C(2)=نام، E(4)=کل کارکرد، G(6)=تعداد شب کاری،
    H(7)=ساعت اضافه‌کاری، I(8)=ساعت جمعه‌کاری، J(9)=مرخصی استفاده‌شده،
    K(10)=مرخصی استعلاجی شرکتی، L(11)=مرخصی استعلاجی تامین‌اجتماعی،
    M(12)=مرخصی بدون حقوق، N(13)=مرخصی تشویقی، O(14)=غیبت، P(15)=کسر کار،
    Q(16)=ماموریت روزانه، S(18)=واحد، V(21)=مانده مرخصی تا پایان ماه

تعداد سطرهای سرستون به‌صورت خودکار تشخیص داده می‌شود (نیازی به ورودی دستی
از کاربر نیست): از ابتدای فایل خط‌به‌خط جلو می‌رویم تا اولین سطری که ستون
«کد پرسنلی» آن واقعاً یک عدد باشد (نه متن سرستون مثل «کد پرسنلی») — همان‌جا
اولین سطر داده واقعی است.

مقادیر عددی دقیقاً طبق Number Format خودِ سلول اکسل قالب‌بندی می‌شوند (نه
مقدار خام اعشاری داخلی) — یعنی همان چیزی که در اکسل روی صفحه دیده می‌شود،
نه رقم دقیق ذخیره‌شده که گاهی به‌خاطر محاسبات داخلی اکسل نویز اعشاری دارد.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

import openpyxl

from app.services.payroll_common import PayrollParseError

# نگاشت فیلد -> شماره ستون (۰-پایه) — دقیقاً همان COLS ابزار HTML مرجع
_COLUMNS: dict[str, int] = {
    "code": 1,
    "name": 2,
    "totalWork": 4,
    "nightDays": 6,
    "overtime": 7,
    "fridayHours": 8,
    "leaveUsed": 9,
    "sickLeave": 10,
    "socialSick": 11,
    "unpaidLeave": 12,
    "bonusLeave": 13,
    "absence": 14,
    "deduction": 15,
    "dailyMission": 16,
    "unit": 18,
    "remainLeave": 21,
}

# ترتیب و برچسب فارسی نمایش هر فیلد — همان ترتیب کارت مرجع
_FIELD_LABELS: list[tuple[str, str]] = [
    ("name", "نام و نام خانوادگی"),
    ("code", "کد پرسنلی"),
    ("totalWork", "کل کارکرد"),
    ("nightDays", "تعداد شب کاری"),
    ("overtime", "ساعت اضافه کاری"),
    ("fridayHours", "ساعت جمعه کاری"),
    ("leaveUsed", "مرخصی استفاده شده"),
    ("sickLeave", "مرخصی استعلاجی شرکتی"),
    ("socialSick", "مرخصی استعلاجی تامین اجتماعی"),
    ("unpaidLeave", "مرخصی بدون حقوق"),
    ("bonusLeave", "مرخصی تشویقی"),
    ("absence", "غیبت"),
    ("deduction", "کسر کار"),
    ("dailyMission", "ماموریت روزانه"),
    ("unit", "واحد"),
    ("remainLeave", "مانده مرخصی تا پایان ماه"),
]

# این دو کلید همیشه رشته‌ای هستند (کد پرسنلی/واحد ممکن است رشته باشد ولی
# نباید طبق Number Format عددی قالب‌بندی شوند)
_TEXT_FIELDS = {"unit"}

_HEADER_SCAN_LIMIT = 30  # حداکثر تا این سطر دنبال اولین ردیف داده می‌گردیم
_FALLBACK_HEADER_ROWS = 4  # اگر تشخیص خودکار جواب نداد


@dataclass
class AttendanceCardItem:
    code: str | None
    fields: list[dict] = field(default_factory=list)  # [{"label": ..., "value": ...}]


def _looks_like_personnel_code(value) -> bool:
    """کد پرسنلی همیشه عدد است — برخلاف متنِ سرستون (مثل «کد پرسنلی»)."""
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        return value.strip().isdigit()
    return False


def _detect_first_data_row(ws) -> int:
    code_col = _COLUMNS["code"]
    for row in ws.iter_rows(min_row=1, max_row=_HEADER_SCAN_LIMIT):
        cell = row[code_col] if code_col < len(row) else None
        if cell is not None and _looks_like_personnel_code(cell.value):
            return cell.row
    return _FALLBACK_HEADER_ROWS + 1


def _format_number_by_excel_format(value: float, number_format: str | None) -> str:
    """
    مقدار عددی را دقیقاً طبق Number Format سلول اکسل قالب‌بندی می‌کند — همان
    چیزی که در خودِ اکسل روی صفحه دیده می‌شود (مثلاً فرمت «0.0» یعنی همیشه
    دقیقاً یک رقم اعشار، حتی اگر مقدار خام 0.999999999989 باشد → «1.0»).
    فقط رایج‌ترین الگوها (تعداد رقم اعشار، جداکننده هزارگان، درصد) پشتیبانی
    می‌شود؛ برای فرمت‌های خیلی خاص (تاریخ و...) یا «General»، فقط نویز
    اعشاری بسیار ریز حذف می‌شود.
    """
    fmt = number_format or "General"

    if fmt in ("General", "@"):
        rounded = round(value, 6)
        if rounded == int(rounded):
            return str(int(rounded))
        return f"{rounded:.6f}".rstrip("0").rstrip(".")

    decimals = 0
    match = re.search(r"\.([0#]+)", fmt)
    if match:
        decimals = len(match.group(1))

    if "%" in fmt:
        return f"{round(value * 100, decimals):.{decimals}f}%"

    rounded = round(value, decimals)
    if "#,##0" in fmt or "#,##" in fmt:
        return f"{rounded:,.{decimals}f}"
    return f"{rounded:.{decimals}f}"


def _format_cell(cell, is_text_field: bool) -> str:
    if cell is None or cell.value is None:
        return ""
    value = cell.value
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)) and not is_text_field:
        return _format_number_by_excel_format(float(value), cell.number_format)
    return str(value).strip()


def parse_attendance_cards_xlsx(file_bytes: bytes) -> list[AttendanceCardItem]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:  # noqa: BLE001 - هر خطای خواندن فایل به پیام قابل‌نمایش تبدیل می‌شود
        raise PayrollParseError("فایل اکسل قابل‌خواندن نیست — فرمت آن را بررسی کنید.") from e

    ws = wb.worksheets[0]
    first_data_row = _detect_first_data_row(ws)
    items: list[AttendanceCardItem] = []

    for row in ws.iter_rows(min_row=first_data_row):
        raw = {}
        for field_name, col_idx in _COLUMNS.items():
            cell = row[col_idx] if col_idx < len(row) else None
            raw[field_name] = _format_cell(cell, field_name in _TEXT_FIELDS)

        code = raw.get("code") or None
        name = raw.get("name") or ""
        if not code and not name:
            continue  # ردیف خالی/تزئینی

        fields = [{"label": label, "value": raw.get(key, "") or "—"} for key, label in _FIELD_LABELS]
        items.append(AttendanceCardItem(code=code, fields=fields))

    return items
