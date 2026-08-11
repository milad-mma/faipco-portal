"""
پارس فایل XLSX فیش حقوقی.

ساختار واقعی این فایل‌ها (خروجی مستقیم "چاپ به Excel" همان گزارش SSRS): هر
پرسنل یک بلوک از سطرها را اشغال می‌کند که با یک سطر «مشخصات» (کد پرسنلی/نام/
مرکز هزینه) شروع می‌شود. زیرِ آن، یک سطر «سرستون» چهار Section را با نامشان
(«وام»، «کسور»، «مزایا»، «سایر») در چهار محدوده‌ی ستونی جدا مشخص می‌کند؛ در
سطرهای بعدی، هر Section مستقل از بقیه، ردیف‌های خودش را به شکل (مقدار در
ستون کوچک‌تر، برچسب در ستون بزرگ‌تر همان سطر) دارد.

چون این فرمت هیچ Header ستونی به سبک جدول معمولی ندارد (مختصات کاملاً شبیه
همان چیدمان بصری گزارش/PDF است، نه یک Export تخت)، پارسر کاملاً Positional و
Generic عمل می‌کند — نه به نام فایل وابسته است و نه به تعداد/ترتیب دقیق
Section ها (اگر Section ناشناخته‌ای هم باشد، به همان اسمش اضافه می‌شود).

الگوریتم:
  1. هر سلولی که دقیقاً برابر «کد پرسنلی:» باشد، شروع یک بلوک پرسنل جدید است.
  2. همان سطر (سطر مشخصات) با یک عبور «مقدار سپس برچسب» (مثل XML) به
     (کد پرسنلی، نام، مرکز هزینه، ...) شکسته می‌شود.
  3. نزدیک‌ترین سطر زیرِ آن که شامل نام‌های شناخته‌شده («وام»/«کسور»/«مزایا»/
     «سایر») باشد، ستون شروع هر Section را مشخص می‌کند.
  4. برای هر سطر، سلول‌های هر Section (بر اساس محدوده ستونی‌اش) به‌صورت
     (اولین سلول غیرخالی = مقدار، آخرین سلول غیرخالی = برچسب) خوانده
     می‌شوند — به‌جز ردیف‌هایی که برچسبشان با «جمع» شروع می‌شود (این‌ها
     ویجت‌های جمع‌بندی پایین فیش‌اند که به‌صورت تصادفی در محدوده ستونی یک
     Section افتاده‌اند، نه یک قلم واقعی از آن Section — به فوتر منتقل
     می‌شوند).
  5. نوار جمع‌بندی پایین فیش (که در XLSX هم مثل XML، برچسب و مقدارش همیشه
     مجاور هم نیستند) با خوشه‌بندی سطری داخل هر گروه ستونی استخراج می‌شود
     (تفصیل در _extract_footer_rows).
"""
from __future__ import annotations

import io
import re

import openpyxl

from app.services.payroll_common import FOOTER_LABEL_COLUMN, ParsedReceiptItem, PayrollParseError, ReceiptSection

_INFO_LABEL_CODE = "کد پرسنلی:"
_PERIOD_LABEL = "فیش حقوق ماه"  # همیشه چند سطر قبل از «کد پرسنلی:» همان بلوک می‌آید — برای تشخیص دقیق مرز واقعی بلوک
# چند برچسب شناخته‌شده که برخلاف «کد پرسنلی:»/«نام و نام خانوادگی:» با «:»
# ختم نمی‌شوند ولی هنوز هم به‌وضوح «برچسب» هستند نه «مقدار» (اصطلاحات عمومی
# گزارش حقوق و دستمزد، نه داده اختصاصی یک سازمان خاص)
_KNOWN_LABEL_WORDS_NO_COLON = {"سال", "فیش حقوق ماه", "ماه"}
_KNOWN_SECTION_NAMES = {"وام", "کسور", "مزایا", "سایر"}


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _row_cells(ws, row: int, max_col: int) -> list[tuple[int, str]]:
    """[(شماره ستون, متن)] سلول‌های غیرخالی یک سطر، به ترتیب ستون."""
    out = []
    for c in range(1, max_col + 1):
        v = _cell_str(ws.cell(row=row, column=c).value)
        if v:
            out.append((c, v))
    return out


def _is_label_text(text: str) -> bool:
    return text.endswith(":") or text.endswith("：") or text in _KNOWN_LABEL_WORDS_NO_COLON


def _pair_row_value_then_label(cells: list[tuple[int, str]]) -> list[dict]:
    """
    مثل payroll_xml._pair_stream ولی روی سلول‌های یک سطر Excel: هر سلولی که
    برچسب شناخته شود، بلافاصله سلول غیرخالی ماقبلش (در همان لیست) «مقدار»
    همان برچسب می‌شود.
    """
    rows: list[dict] = []
    i = 0
    n = len(cells)
    while i < n:
        col, text = cells[i]
        if _is_label_text(text):
            rows.append({"label": text.rstrip(": ："), "value": ""})
            i += 1
            continue
        if i + 1 < n and _is_label_text(cells[i + 1][1]):
            rows.append({"label": cells[i + 1][1].rstrip(": ："), "value": text})
            i += 2
            continue
        rows.append({"label": f"col{col}", "value": text})
        i += 1
    return [r for r in rows if r["value"].strip()]


def _find_code(rows: list[dict]) -> str | None:
    for row in rows:
        if "کد پرسنلی" in row["label"]:
            return row["value"]
    return None


def _extract_footer_rows(
    ws, footer_start_row: int, block_end: int, max_col: int, ranges: list[tuple[int, int, str]]
) -> list[dict]:
    """
    نوار جمع‌بندی پایین فیش را استخراج می‌کند. چون محل دقیق برچسب/مقدار هر
    مورد نسبت به هم یکنواخت نیست (گاهی هم‌سطر، گاهی چند سطر فاصله)، ابتدا هر
    سلول را بر اساس ستونش به یکی از همان ۴ گروه اصلی (وام/کسور/مزایا/سایر)
    نسبت می‌دهد، سپس فقط داخل همان گروه (نه کل فوتر) دنبال نزدیک‌ترین همسایه
    می‌گردد — این کار احتمال قاطی‌شدن مقدارِ یک ستون با برچسبِ ستون دیگر را
    از بین می‌برد.
    """

    def which_column(col: int) -> str | None:
        for start, end, name in ranges:
            if start <= col <= end:
                return name
        return None

    rows_map: dict[int, list[tuple[int, str, str | None]]] = {}
    for r in range(footer_start_row, block_end + 1):
        for c in range(1, max_col + 1):
            text = _cell_str(ws.cell(row=r, column=c).value)
            if text:
                rows_map.setdefault(r, []).append((c, text, which_column(c)))

    results: list[dict] = []
    used_rows: set[int] = set()
    sorted_rows = sorted(rows_map.keys())

    # مرحله ۱: سطرهایی که خودشان هم برچسب شناخته‌شده و هم یک مقدار دارند
    for r in sorted_rows:
        items = rows_map[r]
        labels_here = [(c, t, cn) for c, t, cn in items if t in FOOTER_LABEL_COLUMN]
        values_here = [(c, t, cn) for c, t, cn in items if t not in FOOTER_LABEL_COLUMN]
        if labels_here and values_here:
            for c, t, _cn in labels_here:
                nearest = min(values_here, key=lambda v: abs(v[0] - c))
                results.append({"label": t, "value": nearest[1], "column": FOOTER_LABEL_COLUMN[t]})
            used_rows.add(r)

    # مرحله ۲: سطرهای «تک‌برچسب» (بدون مقدار همان سطر) را با نزدیک‌ترین سطرِ
    # «تک‌مقدار» در همان گروه ستونی (حداکثر ۳ سطر فاصله) جفت می‌کند
    remaining_labels: list[tuple[int, str, str]] = []  # (row, text, column)
    remaining_values: list[tuple[int, str, str | None]] = []  # (row, text, column)
    for r in sorted_rows:
        if r in used_rows:
            continue
        items = rows_map[r]
        labels_here = [(c, t, cn) for c, t, cn in items if t in FOOTER_LABEL_COLUMN]
        values_here = [(c, t, cn) for c, t, cn in items if t not in FOOTER_LABEL_COLUMN]
        if labels_here and not values_here:
            for _c, t, cn in labels_here:
                remaining_labels.append((r, t, cn))
        elif values_here and not labels_here:
            for _c, t, cn in values_here:
                remaining_values.append((r, t, cn))

    claimed: set[int] = set()
    for lr, lt, lcn in remaining_labels:
        best_idx, best_dist = None, None
        for idx, (vr, vt, vcn) in enumerate(remaining_values):
            if idx in claimed or vcn != lcn:
                continue
            dist = abs(vr - lr)
            if dist <= 3 and (best_dist is None or dist < best_dist):
                best_idx, best_dist = idx, dist
        if best_idx is not None:
            claimed.add(best_idx)
            results.append({"label": lt, "value": remaining_values[best_idx][1], "column": FOOTER_LABEL_COLUMN[lt]})

    return results


def parse_salary_receipt_items_xlsx(file_bytes: bytes) -> list[ParsedReceiptItem]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:  # noqa: BLE001 - فایل خراب/فرمت نامعتبر
        raise PayrollParseError(f"فایل XLSX معتبر نیست: {e}") from e

    ws = wb.worksheets[0]
    max_col = ws.max_column
    max_row = ws.max_row

    # ---------- ۱. پیدا کردن شروع هر بلوک پرسنل ----------
    info_rows: list[int] = []
    period_rows: list[int] = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            text = _cell_str(ws.cell(row=r, column=c).value)
            if text == _INFO_LABEL_CODE:
                info_rows.append(r)
            elif text == _PERIOD_LABEL:
                period_rows.append(r)

    if not info_rows:
        raise PayrollParseError(
            f"هیچ سطر «{_INFO_LABEL_CODE}» در فایل پیدا نشد — این فایل با ساختار مورد انتظار مطابقت ندارد"
        )

    # ---------- عنوان بالای فیش (مثلاً «Faipco») — یک‌بار، قبل از اولین بلوک ----------
    report_title: str | None = None
    first_block_hint = min(period_rows) if period_rows else min(info_rows)
    for r in range(1, first_block_hint):
        for c in range(1, max_col + 1):
            text = _cell_str(ws.cell(row=r, column=c).value)
            if text:
                report_title = text
                break
        if report_title:
            break

    def _true_block_start(info_row: int) -> int:
        candidates = [p for p in period_rows if p <= info_row and info_row - p <= 5]
        return min(candidates) if candidates else info_row

    block_start_rows = sorted({_true_block_start(r) for r in info_rows})
    if len(block_start_rows) < len(info_rows):
        block_start_rows = sorted(set(block_start_rows) | set(info_rows))

    items: list[ParsedReceiptItem] = []

    for idx, block_start in enumerate(block_start_rows):
        block_end = (block_start_rows[idx + 1] - 1) if idx + 1 < len(block_start_rows) else max_row
        info_row = next((r for r in info_rows if block_start <= r <= block_end), block_start)

        # ---------- ۲. مشخصات فیش ----------
        header_rows: list[dict] = []
        for r in range(block_start, info_row + 1):
            header_rows.extend(_pair_row_value_then_label(_row_cells(ws, r, max_col)))
        code = _find_code(header_rows)

        # ---------- ۳. سطر سرستون Section ها ----------
        section_header_row: int | None = None
        section_starts: list[tuple[int, str]] = []
        for r in range(info_row + 1, block_end + 1):
            found = [
                (c, _cell_str(ws.cell(row=r, column=c).value))
                for c in range(1, max_col + 1)
                if _cell_str(ws.cell(row=r, column=c).value) in _KNOWN_SECTION_NAMES
            ]
            if len(found) >= 2:
                section_header_row = r
                section_starts = found
                break

        sections: list[ReceiptSection] = []
        data_start_row = (section_header_row + 1) if section_header_row else (info_row + 1)
        footer_start_row = block_end + 1
        ranges: list[tuple[int, int, str]] = []

        if section_starts:
            section_starts.sort(key=lambda x: x[0])
            for i, (start_col, name) in enumerate(section_starts):
                end_col = section_starts[i + 1][0] - 1 if i + 1 < len(section_starts) else max_col
                ranges.append((start_col, end_col, name))

            section_rows_map: dict[str, list[dict]] = {name: [] for _, _, name in ranges}
            last_data_row = section_header_row

            for r in range(data_start_row, block_end + 1):
                matched_this_row = False
                for start_col, end_col, name in ranges:
                    cells = [
                        (c, _cell_str(ws.cell(row=r, column=c).value))
                        for c in range(start_col, end_col + 1)
                        if _cell_str(ws.cell(row=r, column=c).value)
                    ]
                    if not cells or len(cells) == 1:
                        continue
                    value_col, value_text = cells[0]
                    label_col, label_text = cells[-1]
                    if not (value_text and label_text and value_col != label_col):
                        continue
                    if name != "سایر" and (label_text.startswith("جمع") or label_text in FOOTER_LABEL_COLUMN):
                        # این یک ویجت جمع‌بندی پایین فیش است که به‌صورت تصادفی
                        # در محدوده ستونی این Section افتاده — قلم واقعی این
                        # Section نیست؛ به فوتر (مرحله بعد) واگذار می‌شود.
                        # نکته مهم: این استثنا فقط برای وام/کسور/مزایا اعمال
                        # می‌شود، نه «سایر» — چون در «سایر» (بخش Attendance در
                        # XML)، آیتم‌هایی مثل «جمع مزایا»، «جمع کسور»، «خالص
                        # پرداختی» و «جمع ماههای کارکرد» قلم‌های واقعی و
                        # قانونی همان بخش‌اند، نه نشتی از فوتر.
                        continue
                    section_rows_map[name].append({"label": label_text, "value": value_text})
                    matched_this_row = True
                if matched_this_row:
                    last_data_row = r

            for start_col, end_col, name in ranges:
                if section_rows_map[name]:
                    sections.append(ReceiptSection(title=name, rows=section_rows_map[name]))

            footer_start_row = last_data_row + 1

        # ---------- ۴. جمع‌بندی پایین بلوک ----------
        footer_rows = (
            _extract_footer_rows(ws, footer_start_row, block_end, max_col, ranges) if ranges else []
        )

        items.append(
            ParsedReceiptItem(
                code=code,
                report_title=report_title,
                header_rows=header_rows,
                sections=sections,
                footer_rows=footer_rows,
            )
        )

    return items
